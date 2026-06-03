"""
sr_minds_formatter.py — Mindsフォーマット エビデンステーブル + SoFテーブル生成

Usage:
    python sr_minds_formatter.py \
        --input ./sr_output/extracted.csv \
        --output ./sr_output/minds_evidence_table.xlsx \
        --pico-q "がんサバイバーへの運動介入はQOLを改善するか？"
"""

import argparse
from datetime import date

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ---- カラーパレット ----
COLOR_HEADER = "1F3864"      # 濃紺
COLOR_SUBHEADER = "2E75B6"   # 青
COLOR_LOW = "C6EFCE"         # 緑（Low risk）
COLOR_SOME = "FFEB9C"        # 黄（Some concerns）
COLOR_HIGH = "FFC7CE"        # 赤（High risk）
COLOR_GRADE_HIGH = "375623"
COLOR_GRADE_MOD = "538135"
COLOR_GRADE_LOW2 = "C6EFCE"
COLOR_GRADE_VLOW = "FFEB9C"

ROB_COLUMNS = [
    "rob_randomization",
    "rob_allocation",
    "rob_blinding",
    "rob_attrition",
    "rob_reporting",
    "rob_overall",
]

ROB_LABELS = {
    "rob_randomization": "無作為化",
    "rob_allocation": "割付け隠蔽",
    "rob_blinding": "盲検化",
    "rob_attrition": "脱落",
    "rob_reporting": "選択的報告",
    "rob_overall": "総合",
}

GRADE_SYMBOLS = {
    "High": "⊕⊕⊕⊕",
    "Moderate": "⊕⊕⊕◯",
    "Low": "⊕⊕◯◯",
    "Very Low": "⊕◯◯◯",
}


def thin_border():
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def header_cell(ws, row, col, value, bg_color=COLOR_HEADER, font_color="FFFFFF", bold=True, wrap=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = PatternFill("solid", fgColor=bg_color)
    cell.font = Font(bold=bold, color=font_color, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    cell.border = thin_border()
    return cell


def data_cell(ws, row, col, value, wrap=True, bg_color=None, bold=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=wrap)
    cell.border = thin_border()
    cell.font = Font(size=9, bold=bold)
    if bg_color:
        cell.fill = PatternFill("solid", fgColor=bg_color)
    return cell


def rob_color(value: str) -> str | None:
    v = str(value).lower()
    if "low" in v:
        return COLOR_LOW
    if "some" in v or "concern" in v:
        return COLOR_SOME
    if "high" in v:
        return COLOR_HIGH
    return None


def infer_grade(rob_overall_values: list[str]) -> str:
    """RoB overall判定からGRADE確実性を簡易推定。"""
    highs = sum(1 for v in rob_overall_values if "high" in str(v).lower())
    somes = sum(1 for v in rob_overall_values if "some" in str(v).lower())
    n = len(rob_overall_values)
    if n == 0:
        return "Very Low"
    high_ratio = highs / n
    if high_ratio > 0.5:
        return "Very Low"
    if high_ratio > 0.2 or somes / n > 0.5:
        return "Low"
    if somes / n > 0.2:
        return "Moderate"
    return "High"


# ---- Sheet 1: エビデンステーブル ----

EVIDENCE_COLS = [
    ("著者・年", "Authors", 20),
    ("対象\n(P)", "population", 30),
    ("介入\n(I)", "intervention", 30),
    ("対照\n(C)", "comparison", 20),
    ("アウトカム\n(O)", "outcomes", 35),
    ("追跡期間", "follow_up", 12),
    ("研究デザイン", "study_design", 14),
    ("無作為化", "rob_randomization", 12),
    ("割付け隠蔽", "rob_allocation", 12),
    ("盲検化", "rob_blinding", 12),
    ("脱落", "rob_attrition", 12),
    ("選択的報告", "rob_reporting", 12),
    ("RoB総合", "rob_overall", 12),
    ("特記事項", "notes", 30),
    ("URL", "URL", 30),
]


def write_evidence_sheet(ws, df: pd.DataFrame, pico_q: str):
    ws.title = "エビデンステーブル"
    ws.freeze_panes = "A3"

    # タイトル行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(EVIDENCE_COLS))
    title_cell = ws.cell(row=1, column=1, value=f"エビデンステーブル — {pico_q}")
    title_cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
    title_cell.font = Font(bold=True, color="FFFFFF", size=12)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # ヘッダー行
    for col_idx, (label, _, width) in enumerate(EVIDENCE_COLS, start=1):
        header_cell(ws, 2, col_idx, label, bg_color=COLOR_SUBHEADER)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 36

    # データ行
    for row_idx, row in enumerate(df.to_dict("records"), start=3):
        author_year = f"{str(row.get('Authors', '')).split(',')[0]} et al., {row.get('Year', '')}"

        for col_idx, (_, field, _) in enumerate(EVIDENCE_COLS, start=1):
            if field == "Authors":
                val = author_year
            else:
                val = row.get(field, "")

            bg = rob_color(val) if field in ROB_COLUMNS else None
            data_cell(ws, row_idx, col_idx, val, bg_color=bg)

        ws.row_dimensions[row_idx].height = 60


# ---- Sheet 2: SoFテーブル（GRADE Summary of Findings）----

SOF_OUTCOME_COLS = [
    ("アウトカム", 25),
    ("研究数\n(参加者数)", 14),
    ("研究デザイン", 14),
    ("RoB", 12),
    ("非直接性", 12),
    ("非一貫性", 12),
    ("不精確さ", 12),
    ("出版バイアス", 14),
    ("上昇要因", 12),
    ("エビデンスの確実性", 20),
    ("効果の要約", 35),
]


def write_sof_sheet(ws, df: pd.DataFrame, pico_q: str, outcomes_list: list[str]):
    ws.title = "SoFテーブル(GRADE)"
    ws.freeze_panes = "A3"

    # タイトル行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(SOF_OUTCOME_COLS))
    tc = ws.cell(row=1, column=1, value=f"Summary of Findings — {pico_q}")
    tc.fill = PatternFill("solid", fgColor=COLOR_HEADER)
    tc.font = Font(bold=True, color="FFFFFF", size=12)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # ヘッダー
    for col_idx, (label, width) in enumerate(SOF_OUTCOME_COLS, start=1):
        header_cell(ws, 2, col_idx, label, bg_color=COLOR_SUBHEADER)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 40

    n_studies = len(df)
    n_participants = "不明"  # 抄録からの自動取得は困難なため手動記入欄

    rob_values = df["rob_overall"].tolist() if "rob_overall" in df.columns else []
    grade = infer_grade(rob_values)
    grade_symbol = GRADE_SYMBOLS.get(grade, "⊕◯◯◯")

    grade_bg = {
        "High": COLOR_GRADE_HIGH,
        "Moderate": COLOR_GRADE_MOD,
        "Low": COLOR_GRADE_LOW2,
        "Very Low": COLOR_GRADE_VLOW,
    }.get(grade, COLOR_GRADE_VLOW)

    for row_idx, outcome in enumerate(outcomes_list, start=3):
        values = [
            outcome,
            f"{n_studies}件\n({n_participants}名)",
            "RCT",
            "（RoBシートを参照）",
            "なし",
            "—",
            "—",
            "なし",
            "なし",
            f"{grade_symbol}\n{grade}",
            "（手動記入）",
        ]
        for col_idx, val in enumerate(values, start=1):
            bg = grade_bg if col_idx == 10 else None
            font_color = "FFFFFF" if col_idx == 10 and grade in ("High", "Moderate") else "000000"
            c = data_cell(ws, row_idx, col_idx, val, bg_color=bg)
            if col_idx == 10:
                c.font = Font(bold=True, size=9, color=font_color)
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[row_idx].height = 50

    # 凡例
    legend_row = len(outcomes_list) + 4
    ws.merge_cells(
        start_row=legend_row, start_column=1, end_row=legend_row, end_column=len(SOF_OUTCOME_COLS)
    )
    legend = ws.cell(
        row=legend_row,
        column=1,
        value="凡例: ⊕⊕⊕⊕ High（非常に強い）/ ⊕⊕⊕◯ Moderate / ⊕⊕◯◯ Low / ⊕◯◯◯ Very Low（非常に弱い）",
    )
    legend.font = Font(italic=True, size=9)
    legend.alignment = Alignment(horizontal="left")


# ---- Sheet 3: バイアスリスク一覧 ----

def write_rob_sheet(ws, df: pd.DataFrame):
    ws.title = "バイアスリスク"
    ws.freeze_panes = "A3"

    cols = [("著者・年", 25)] + [(v, 16) for v in ROB_LABELS.values()]
    for col_idx, (label, width) in enumerate(cols, start=1):
        header_cell(ws, 1, col_idx, label, bg_color=COLOR_SUBHEADER)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30

    for row_idx, row in enumerate(df.to_dict("records"), start=2):
        author_year = f"{str(row.get('Authors', '')).split(',')[0]} et al., {row.get('Year', '')}"
        data_cell(ws, row_idx, 1, author_year)
        for col_idx, rob_key in enumerate(ROB_COLUMNS, start=2):
            val = row.get(rob_key, "")
            data_cell(ws, row_idx, col_idx, val, bg_color=rob_color(val))
        ws.row_dimensions[row_idx].height = 30


# ---- メイン ----

def generate_minds_table(input_csv: str, output_xlsx: str, pico_q: str, outcomes: str):
    df = pd.read_csv(input_csv).fillna("")
    outcomes_list = [o.strip() for o in outcomes.split(",") if o.strip()]

    wb = Workbook()
    ws1 = wb.active
    ws2 = wb.create_sheet()
    ws3 = wb.create_sheet()

    write_evidence_sheet(ws1, df, pico_q)
    write_sof_sheet(ws2, df, pico_q, outcomes_list)
    write_rob_sheet(ws3, df)

    wb.save(output_xlsx)
    print(f"[完了] {output_xlsx} を出力しました（シート: エビデンステーブル / SoFテーブル / バイアスリスク）")


def main():
    parser = argparse.ArgumentParser(description="Mindsエビデンステーブル生成")
    parser.add_argument("--input", required=True, help="抽出済みCSV（sr_data_extraction.py出力）")
    parser.add_argument("--output", default="minds_evidence_table.xlsx")
    parser.add_argument("--pico-q", default="介入はアウトカムを改善するか？", help="レビューの問い")
    parser.add_argument("--outcomes", default="QOL, 疲労, 身体機能", help="SoFテーブルに記載するアウトカム（カンマ区切り）")
    args = parser.parse_args()

    generate_minds_table(args.input, args.output, args.pico_q, args.outcomes)


if __name__ == "__main__":
    main()
